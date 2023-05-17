import re
import requests
import pandas as pd
import openpyxl
from lxml import html
urls = [
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/large-cap-fund.html", "Large Cap"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/large-and-mid-cap-fund.html", "Large and Mid cap"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/elss.html", "ELSS"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/focused-fund.html", "Focused"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/mid-cap-fund.html", "Mid Cap"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/aggressive-hybrid-fund.html", "Aggressive Hybrid"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/conservative-hybrid-fund.html", "Conservative Hybrid"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/equity-savings.html", "Equity Savings"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/dynamic-asset-allocation-or-balanced-advantage.html", "Dynamic Asset"),
    ("https://www.moneycontrol.com/mutual-funds/performance-tracker/returns/multi-cap-fund.html", "Multicap"),
    ("https://www.moneycontrol.com/mutual-funds/best-funds/hybrid/returns/1", "Hybrid"),
    ("https://www.moneycontrol.com/mutual-funds/best-funds/debt/returns/1", "Debt funds")
]
headers_dict = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
}
wb = openpyxl.Workbook()

default_sheet = wb["Sheet"]
wb.remove(default_sheet)
for url, sheet_name in urls:
    response = requests.get(url, headers=headers_dict)
    html_source = response.text
    html_source = re.sub(r'<.*?>', lambda g: g.group(0).upper(), html_source)
    tree = html.fromstring(html_source)
    tables = tree.xpath('//table')
    dataframes = pd.read_html(html_source, header=None)
    ws = wb.create_sheet(sheet_name)
    excel_headers = ["Scheme Name", "Plan", "Category Name", "Crisil Rank", "AuM (Cr)", "1W", "1M", "3M", "6M", "YTD", "1Y", "2Y", "3Y", "5Y", "10Y"]
    for col_num, header in enumerate(excel_headers):
        ws.cell(row=1, column=col_num + 1).value = header
    
    start_row = 2
    for i, df in enumerate(dataframes):
        for row_num, row_data in df.iterrows():
            for col_num, cell_data in enumerate(row_data):
                if col_num in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
                    cell_value = str(cell_data).strip('%') if isinstance(cell_data, str) else cell_data
                    ws.cell(row=start_row + row_num, column=col_num + 1).value = float(cell_value) / 100 if cell_value != '-' else None
                else:
                    ws.cell(row=start_row + row_num, column=col_num + 1).value = cell_data
        start_row += len(df.index) + 1  # add 1 for the blank row after each table
# create a new worksheet named "Best funds"
best_funds_ws = wb.create_sheet("Best Mutual funds")

# set headers for Best funds worksheet
best_funds_ws.cell(row=1, column=1).value = "Scheme Name"
best_funds_ws.cell(row=1, column=2).value = "3Y"
best_funds_ws.cell(row=1, column=3).value = "Crisil Rank"
best_funds_ws.cell(row=1, column=4).value = "Source"

# iterate over each sheet and extract data
for url, sheet_name in urls:
    response = requests.get(url, headers=headers_dict)
    html_source = response.text
    html_source = re.sub(r'<.*?>', lambda g: g.group(0).upper(), html_source)
    tree = html.fromstring(html_source)
    tables = tree.xpath('//table')
    dataframes = pd.read_html(html_source, header=None)
    
    # iterate over each row of the current sheet
    for df in dataframes:
        for row_data in df.values:
            scheme_name, _, category_name, crisil_rank, _, _, _, _, _, _, _, y2, y3, *_= row_data

            # check if the fund meets the criteria
            if y3 is not None and "%" in y3 and str(crisil_rank) in ["1", "2"] and float(y3.strip('%')) > 20 and "Sponsored" not in scheme_name:
                # add the fund to the Best funds worksheet
                best_funds_ws.append([scheme_name, y3, crisil_rank, sheet_name])
for column_cells in best_funds_ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    best_funds_ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                    
wb.save("MutualFunds.xlsx")

